from __future__ import annotations

import csv
import json
import os
import re
import shutil
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin, urlparse

import requests
import shapefile
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from tqdm import tqdm
from urllib3.util.retry import Retry


API_ROOT = "https://api.planet.com/basemaps/v1"
BASE_DIR = Path(__file__).resolve().parent.parent

QUAD_ID_PATTERN = re.compile(r"^(?P<x>\d+)-(?P<y>\d+)$")
TIFF_NAME_PATTERN = re.compile(
    r"^(?P<x>\d+)-(?P<y>\d+)\.(?:tif|tiff)$",
    re.IGNORECASE,
)

load_dotenv(BASE_DIR / ".env")


# ---------------------------------------------------------------------------
# KONFIGURASI
# ---------------------------------------------------------------------------

def env_bool(name: str, default: bool) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return default

    return raw.strip().lower() in {
        "1",
        "true",
        "yes",
        "y",
        "on",
    }


def resolve_project_path(
    raw_value: str,
    default_value: str,
) -> Path:
    value = (raw_value or default_value).strip()
    value = value.strip('"').strip("'")
    value = os.path.expandvars(os.path.expanduser(value))

    path = Path(value)

    if not path.is_absolute():
        path = BASE_DIR / path

    return path.resolve()


def safe_sheet_name(raw_name: str, fallback: str) -> str:
    cleaned = re.sub(
        r"[\[\]:*?/\\]+",
        "_",
        (raw_name or fallback).strip(),
    )
    cleaned = cleaned[:31].strip()
    return cleaned or fallback[:31]


API_KEY = os.getenv("PL_API_KEY", "").strip()
MOSAIC_NAME = os.getenv("MOSAIC_NAME", "").strip()
BBOX = os.getenv("BBOX", "").strip()

OUTPUT_DIR = resolve_project_path(
    os.getenv("OUTPUT_DIR", ""),
    "downloads",
)

MAX_WORKERS = max(
    1,
    int(os.getenv("MAX_WORKERS", "4")),
)

DRY_RUN = env_bool("DRY_RUN", True)
EXPORT_EXCEL = env_bool("EXPORT_EXCEL", True)
BACKUP_EXCEL = env_bool("BACKUP_EXCEL", True)

EXCEL_OUTPUT_DIR_TEXT = os.getenv(
    "EXCEL_OUTPUT_DIR",
    "",
).strip()

LEGACY_EXCEL_PATH_TEXT = os.getenv(
    "EXCEL_PATH",
    "",
).strip()

if EXCEL_OUTPUT_DIR_TEXT:
    EXCEL_OUTPUT_DIR = resolve_project_path(
        EXCEL_OUTPUT_DIR_TEXT,
        "downloads/excel1test",
    )
elif LEGACY_EXCEL_PATH_TEXT:
    EXCEL_OUTPUT_DIR = resolve_project_path(
        LEGACY_EXCEL_PATH_TEXT,
        "downloads/excel1test/rekap.xlsx",
    ).parent
else:
    EXCEL_OUTPUT_DIR = resolve_project_path(
        "",
        "downloads/excel1test",
    )

AUTO_RUN_NAMING = env_bool(
    "AUTO_RUN_NAMING",
    True,
)

RUN_NAME_FORMAT = os.getenv(
    "RUN_NAME_FORMAT",
    "%H_%M_%d_%m_%Y",
).strip() or "%H_%M_%d_%m_%Y"

RUN_NAME_COLLISION_MODE = os.getenv(
    "RUN_NAME_COLLISION_MODE",
    "suffix",
).strip().lower()

REUSE_PREVIOUS_TIFF = env_bool(
    "REUSE_PREVIOUS_TIFF",
    True,
)

REUSE_METHOD = os.getenv(
    "REUSE_METHOD",
    "hardlink",
).strip().lower()

CREATE_RUN_MANIFEST = env_bool(
    "CREATE_RUN_MANIFEST",
    True,
)

SAVE_RUN_LOG = env_bool(
    "SAVE_RUN_LOG",
    True,
)

LAST_RUN_FILE = resolve_project_path(
    os.getenv("LAST_RUN_FILE", ""),
    "config/last_run.json",
)

CREATE_MASTER_MAP = env_bool(
    "CREATE_MASTER_MAP",
    True,
)

MASTER_MAP_SHEET_NAME = safe_sheet_name(
    os.getenv(
        "MASTER_MAP_SHEET_NAME",
        "MASTER_MAP_QUADS",
    ),
    "MASTER_MAP_QUADS",
)

MASTER_QUADS_SOURCE_TEXT = os.getenv(
    "MASTER_QUADS_SOURCE",
    "",
).strip()

MASTER_QUADS_SOURCE = (
    resolve_project_path(
        MASTER_QUADS_SOURCE_TEXT,
        MASTER_QUADS_SOURCE_TEXT,
    )
    if MASTER_QUADS_SOURCE_TEXT
    else None
)

MASTER_QUAD_ID_FIELD = os.getenv(
    "MASTER_QUAD_ID_FIELD",
    "id",
).strip()

MASTER_QUADS_SHEET = os.getenv(
    "MASTER_QUADS_SHEET",
    "",
).strip()

MASTER_SOURCE_REQUIRED = env_bool(
    "MASTER_SOURCE_REQUIRED",
    False,
)


USE_SELECTED_QUADS = env_bool(
    "USE_SELECTED_QUADS",
    False,
)

SELECTED_QUADS_FILE = resolve_project_path(
    os.getenv("SELECTED_QUADS_FILE", ""),
    "config/selected_quads.json",
)


# ---------------------------------------------------------------------------
# RUN SESSION / OUTPUT
# ---------------------------------------------------------------------------

def sanitize_run_component(value: str) -> str:
    cleaned = re.sub(
        r"[^A-Za-z0-9._-]+",
        "_",
        value.strip(),
    )
    cleaned = re.sub(r"_+", "_", cleaned)
    return cleaned.strip("._-") or "planet_mosaic"


def create_run_context(
    started_at: datetime,
) -> dict[str, object]:
    """Membuat satu nama yang dipakai folder TIFF dan workbook."""
    mosaic_component = sanitize_run_component(
        MOSAIC_NAME
    )

    if AUTO_RUN_NAMING:
        time_component = started_at.strftime(
            RUN_NAME_FORMAT
        )
        base_name = sanitize_run_component(
            f"{time_component}_{mosaic_component}"
        )
    else:
        base_name = mosaic_component

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )
    EXCEL_OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    candidate = base_name
    suffix_number = 2

    while (
        (OUTPUT_DIR / candidate).exists()
        or (
            EXCEL_OUTPUT_DIR
            / f"{candidate}.xlsx"
        ).exists()
    ):
        if RUN_NAME_COLLISION_MODE != "suffix":
            raise RuntimeError(
                "Nama run sudah ada dan "
                "RUN_NAME_COLLISION_MODE bukan 'suffix': "
                f"{candidate}"
            )

        candidate = (
            f"{base_name}_{suffix_number:02d}"
        )
        suffix_number += 1

    output_folder = OUTPUT_DIR / candidate
    excel_path = (
        EXCEL_OUTPUT_DIR
        / f"{candidate}.xlsx"
    )

    output_folder.mkdir(
        parents=True,
        exist_ok=False,
    )

    return {
        "run_name": candidate,
        "started_at": started_at,
        "output_folder": output_folder,
        "excel_path": excel_path,
        "manifest_path": (
            output_folder
            / "run_manifest.json"
        ),
        "log_path": (
            output_folder
            / "run.log"
        ),
    }


def folder_matches_current_mosaic(
    folder: Path,
) -> bool:
    """
    Memastikan TIFF reuse berasal dari mosaic yang sama.

    Quad ID sama pada quarter berbeda bukan file citra yang sama.
    """
    manifest_path = (
        folder / "run_manifest.json"
    )

    if manifest_path.exists():
        try:
            payload = json.loads(
                manifest_path.read_text(
                    encoding="utf-8"
                )
            )

            return (
                str(
                    payload.get(
                        "mosaic_name",
                        "",
                    )
                ).strip()
                == MOSAIC_NAME
            )

        except (
            OSError,
            json.JSONDecodeError,
        ):
            pass

    mosaic_component = sanitize_run_component(
        MOSAIC_NAME
    )

    return (
        folder.name == MOSAIC_NAME
        or mosaic_component in folder.name
    )


def build_previous_tiff_index(
    current_output_folder: Path,
) -> dict[str, Path]:
    """
    Mencari TIFF run terdahulu khusus untuk mosaic aktif.

    Contoh:
    - 1706-1018 dari 2026Q2 boleh dipakai ulang untuk run 2026Q2 lain.
    - 1706-1018 dari 2026Q2 tidak boleh dipakai untuk 2025Q3.
    """
    if not REUSE_PREVIOUS_TIFF:
        return {}

    current_resolved = (
        current_output_folder.resolve()
    )
    excel_resolved = (
        EXCEL_OUTPUT_DIR.resolve()
    )

    index: dict[str, Path] = {}
    mtimes: dict[str, float] = {}

    if not OUTPUT_DIR.exists():
        return index

    for folder in OUTPUT_DIR.iterdir():
        if not folder.is_dir():
            continue

        folder_resolved = (
            folder.resolve()
        )

        if folder_resolved == current_resolved:
            continue

        if (
            folder_resolved == excel_resolved
            or excel_resolved
            in folder_resolved.parents
        ):
            continue

        if not folder_matches_current_mosaic(
            folder
        ):
            continue

        for path in folder.rglob("*.tif"):
            if not path.is_file():
                continue

            quad_id = path.stem

            if not QUAD_ID_PATTERN.fullmatch(
                quad_id
            ):
                continue

            try:
                modified = (
                    path.stat().st_mtime
                )
            except OSError:
                continue

            if modified >= mtimes.get(
                quad_id,
                -1.0,
            ):
                index[quad_id] = path
                mtimes[quad_id] = modified

    return index


def reuse_previous_tiff(
    source: Path,
    target: Path,
) -> str:
    """Hardlink bila memungkinkan; fallback otomatis ke copy."""
    target.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    if REUSE_METHOD == "copy":
        shutil.copy2(source, target)
        return "reused_copy"

    try:
        os.link(source, target)
        return "reused_hardlink"
    except OSError:
        shutil.copy2(source, target)
        return "reused_copy"


def load_selection_metadata() -> dict:
    if not SELECTED_QUADS_FILE.exists():
        return {}

    try:
        return json.loads(
            SELECTED_QUADS_FILE.read_text(
                encoding="utf-8"
            )
        )
    except (OSError, json.JSONDecodeError):
        return {}


def write_json_file(
    path: Path,
    payload: dict,
) -> None:
    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    temporary_path = path.with_suffix(
        path.suffix + ".tmp"
    )

    temporary_path.write_text(
        json.dumps(
            payload,
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    temporary_path.replace(path)


def write_run_state(
    run_context: dict[str, object],
    *,
    status: str,
    completed_at: datetime | None = None,
    counts: dict[str, int] | None = None,
    excel_path: Path | None = None,
    error: str = "",
) -> dict:
    selection = load_selection_metadata()
    started_at = run_context["started_at"]
    output_folder = Path(
        run_context["output_folder"]
    )
    manifest_path = Path(
        run_context["manifest_path"]
    )
    log_path = Path(
        run_context["log_path"]
    )

    payload = {
        "schema_version": 1,
        "status": status,
        "run_name": run_context["run_name"],
        "started_at": (
            started_at.isoformat(
                timespec="seconds"
            )
            if isinstance(started_at, datetime)
            else str(started_at)
        ),
        "completed_at": (
            completed_at.isoformat(
                timespec="seconds"
            )
            if completed_at
            else None
        ),
        "mosaic_name": MOSAIC_NAME,
        "area_mode": selection.get(
            "area_mode",
            "",
        ),
        "bbox": selection.get(
            "bbox",
            BBOX,
        ),
        "quad_count": selection.get(
            "quad_count",
            0,
        ),
        "counts": counts or {},
        "output_folder": str(
            output_folder.resolve()
        ),
        "excel_path": (
            str(excel_path.resolve())
            if excel_path
            else str(
                Path(
                    run_context["excel_path"]
                ).resolve()
            )
        ),
        "manifest_path": str(
            manifest_path.resolve()
        ),
        "run_log_path": str(
            log_path.resolve()
        ),
        "save_run_log": SAVE_RUN_LOG,
        "error": error,
    }

    if CREATE_RUN_MANIFEST:
        write_json_file(
            manifest_path,
            payload,
        )

    write_json_file(
        LAST_RUN_FILE,
        payload,
    )

    return payload


# ---------------------------------------------------------------------------
# HTTP / PLANET API
# ---------------------------------------------------------------------------

def create_session(use_auth: bool = True) -> requests.Session:
    session = requests.Session()

    if use_auth:
        session.auth = (API_KEY, "")

    retry = Retry(
        total=6,
        connect=4,
        read=4,
        status=6,
        backoff_factor=1.5,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset({"GET"}),
        raise_on_status=False,
    )

    adapter = HTTPAdapter(
        max_retries=retry,
        pool_connections=MAX_WORKERS,
        pool_maxsize=MAX_WORKERS,
    )

    session.mount("https://", adapter)
    return session


API_SESSION = create_session(use_auth=True)


def get_json(
    url: str,
    params: dict | None = None,
) -> dict:
    response = API_SESSION.get(
        url,
        params=params,
        timeout=(30, 180),
    )

    if response.status_code == 401:
        raise RuntimeError(
            "401 Unauthorized: API key salah atau tidak terbaca."
        )

    if response.status_code == 403:
        raise RuntimeError(
            "403 Forbidden: akun tidak memiliki akses download."
        )

    response.raise_for_status()
    return response.json()


def validate_bbox(value: str) -> str:
    try:
        values = [
            float(item.strip())
            for item in value.split(",")
        ]
    except ValueError as exc:
        raise RuntimeError(
            "BBOX harus berformat "
            "min_lon,min_lat,max_lon,max_lat."
        ) from exc

    if len(values) != 4:
        raise RuntimeError(
            "BBOX harus berisi tepat empat angka."
        )

    min_lon, min_lat, max_lon, max_lat = values

    if min_lon >= max_lon:
        raise RuntimeError(
            "Longitude kiri harus lebih kecil dari kanan."
        )

    if min_lat >= max_lat:
        raise RuntimeError(
            "Latitude bawah harus lebih kecil dari atas."
        )

    return ",".join(str(item) for item in values)


def find_mosaic(
    mosaic_name: str | None = None,
) -> dict:
    target_name = (
        str(mosaic_name).strip()
        if mosaic_name is not None
        else MOSAIC_NAME
    )
    print(f"Mencari mosaic: {target_name}")

    data = get_json(
        f"{API_ROOT}/mosaics/",
        params={"name__is": target_name},
    )

    mosaic = next(
        (
            item
            for item in data.get("mosaics", [])
            if item.get("name") == target_name
        ),
        None,
    )

    if mosaic is None:
        raise RuntimeError(
            "Mosaic tidak ditemukan atau tidak dapat diakses: "
            f"{target_name}"
        )

    return mosaic


def list_quads(mosaic_id: str, bbox: str) -> list[dict]:
    print("Mencari quad berdasarkan BBOX...")

    url: str | None = (
        f"{API_ROOT}/mosaics/{mosaic_id}/quads/"
    )

    params: dict | None = {
        "bbox": bbox,
        "_page_size": 100,
    }

    quads: list[dict] = []

    while url:
        data = get_json(url, params=params)
        params = None

        quads.extend(data.get("items", []))

        next_url = (
            data.get("_links", {})
            .get("_next")
        )

        url = (
            urljoin(url, next_url)
            if next_url
            else None
        )

    unique = {
        str(quad["id"]): quad
        for quad in quads
        if quad.get("id")
    }

    return [
        unique[quad_id]
        for quad_id in sorted(unique)
    ]


def get_download_url(quad: dict) -> str:
    value = (
        quad.get("_links", {})
        .get("download")
    )

    if isinstance(value, dict):
        value = value.get("href")

    if not isinstance(value, str) or not value:
        raise RuntimeError(
            f"Download link quad {quad.get('id')} tidak tersedia."
        )

    return urljoin(f"{API_ROOT}/", value)


def download_quad(
    quad: dict,
    output_folder: Path,
    previous_tiff_index: dict[str, Path],
) -> str:
    quad_id = str(quad["id"])

    target = output_folder / f"{quad_id}.tif"
    temporary = output_folder / f"{quad_id}.tif.part"

    if target.exists() and target.stat().st_size > 0:
        return "existing"

    previous_source = previous_tiff_index.get(
        quad_id
    )

    if (
        previous_source
        and previous_source.exists()
        and previous_source.stat().st_size > 0
    ):
        return reuse_previous_tiff(
            previous_source,
            target,
        )

    download_url = get_download_url(quad)

    use_auth = (
        urlparse(download_url).hostname
        == "api.planet.com"
    )

    session = create_session(use_auth=use_auth)

    with session.get(
        download_url,
        stream=True,
        allow_redirects=True,
        timeout=(30, 900),
    ) as response:
        response.raise_for_status()

        with temporary.open("wb") as stream:
            for chunk in response.iter_content(
                chunk_size=1024 * 1024
            ):
                if chunk:
                    stream.write(chunk)

    temporary.replace(target)
    return "downloaded"


# ---------------------------------------------------------------------------
# QUAD ID / FILE LOKAL
# ---------------------------------------------------------------------------

def parse_quad_id(
    quad_id: str,
) -> tuple[int, int] | None:
    match = QUAD_ID_PATTERN.fullmatch(
        str(quad_id).strip()
    )

    if not match:
        return None

    return (
        int(match.group("x")),
        int(match.group("y")),
    )


def normalize_quad_id(value: object) -> str | None:
    if value is None:
        return None

    text = str(value).strip()

    if not QUAD_ID_PATTERN.fullmatch(text):
        return None

    return text


def load_selected_quad_ids() -> set[str] | None:
    """
    Membaca filter hasil Area Selector.

    None berarti filter dinonaktifkan.
    Set kosong dianggap sebagai konfigurasi yang tidak valid.
    """
    if not USE_SELECTED_QUADS:
        return None

    if not SELECTED_QUADS_FILE.exists():
        raise RuntimeError(
            "USE_SELECTED_QUADS=true tetapi file seleksi tidak ditemukan: "
            f"{SELECTED_QUADS_FILE}"
        )

    try:
        payload = json.loads(
            SELECTED_QUADS_FILE.read_text(encoding="utf-8")
        )
    except json.JSONDecodeError as exc:
        raise RuntimeError(
            f"JSON seleksi tidak valid: {SELECTED_QUADS_FILE}"
        ) from exc

    raw_ids = payload.get("quad_ids", [])

    selected_ids = {
        normalized
        for value in raw_ids
        if (normalized := normalize_quad_id(value))
    }

    if not selected_ids:
        raise RuntimeError(
            "File selected_quads.json tidak berisi Quad ID valid."
        )

    configured_mosaic = str(
        payload.get("mosaic_name", "")
    ).strip()

    if configured_mosaic and configured_mosaic != MOSAIC_NAME:
        print(
            "PERINGATAN: mosaic pada selected_quads.json berbeda "
            f"({configured_mosaic}) dari MOSAIC_NAME ({MOSAIC_NAME})."
        )

    return selected_ids


def extract_period_code(mosaic_name: str) -> str:
    match = re.search(
        r"(?P<year>\d{4})q(?P<quarter>[1-4])",
        mosaic_name,
        re.IGNORECASE,
    )

    if match:
        return (
            f"{match.group('year')}"
            f"Q{match.group('quarter')}"
        )

    cleaned = re.sub(
        r"[^A-Za-z0-9]+",
        "_",
        mosaic_name,
    ).strip("_")

    return cleaned[-20:] or "MOSAIC"


def scan_local_tiffs(
    output_folder: Path,
) -> tuple[dict[str, Path], list[Path]]:
    valid: dict[str, Path] = {}
    invalid: list[Path] = []

    for path in sorted(output_folder.glob("*")):
        if not path.is_file():
            continue

        if path.suffix.lower() not in {".tif", ".tiff"}:
            continue

        match = TIFF_NAME_PATTERN.fullmatch(path.name)

        if not match:
            invalid.append(path)
            continue

        quad_id = (
            f"{match.group('x')}-"
            f"{match.group('y')}"
        )

        valid[quad_id] = path

    return valid, invalid


def build_active_records(
    quads: list[dict],
    download_results: dict[str, dict[str, str]],
    output_folder: Path,
) -> tuple[list[dict], list[Path]]:
    api_by_id = {
        str(quad["id"]): quad
        for quad in quads
        if quad.get("id")
    }

    local_files, invalid_files = scan_local_tiffs(
        output_folder
    )

    # Saat pemilihan area dari UI aktif, laporan periode hanya
    # memuat Quad yang berada pada seleksi aktif. TIFF dari AOI lama
    # tetap tersimpan di disk, tetapi tidak dilabeli "Extra".
    if USE_SELECTED_QUADS:
        all_ids = sorted(
            set(api_by_id)
        )
    else:
        all_ids = sorted(
            set(api_by_id) | set(local_files)
        )

    processed_at = datetime.now()
    records: list[dict] = []

    for quad_id in all_ids:
        parsed = parse_quad_id(quad_id)

        if parsed is None:
            continue

        x_value, y_value = parsed
        quad = api_by_id.get(quad_id, {})
        local_path = local_files.get(quad_id)
        result = download_results.get(quad_id, {})

        error_text = result.get("error", "")
        process_status = result.get("status", "")

        if error_text:
            status = "Failed"
        elif local_path and quad_id not in api_by_id:
            status = "Extra"
        elif local_path and process_status == "downloaded":
            status = "Downloaded"
        elif local_path and process_status.startswith("reused_"):
            status = "Reused"
        elif local_path:
            status = "Existing"
        else:
            status = "Missing"

        size_mb = (
            round(
                local_path.stat().st_size
                / (1024 * 1024),
                2,
            )
            if local_path and local_path.exists()
            else 0
        )

        relative_path = ""

        if local_path:
            try:
                relative_path = str(
                    local_path.resolve().relative_to(
                        BASE_DIR.resolve()
                    )
                )
            except ValueError:
                relative_path = str(local_path.resolve())

        records.append(
            {
                "quad_id": quad_id,
                "x": x_value,
                "y": y_value,
                "mosaic_name": MOSAIC_NAME,
                "available_in_api": quad_id in api_by_id,
                "filename": (
                    local_path.name
                    if local_path
                    else f"{quad_id}.tif"
                ),
                "status": status,
                "coverage": quad.get(
                    "percent_covered",
                    "",
                ),
                "size_mb": size_mb,
                "relative_path": relative_path,
                "absolute_path": (
                    str(local_path.resolve())
                    if local_path
                    else ""
                ),
                "processed_at": processed_at,
                "error": error_text,
            }
        )

    return records, invalid_files


# ---------------------------------------------------------------------------
# SUMBER MASTER QUAD INDONESIA
# ---------------------------------------------------------------------------

def find_case_insensitive_field(
    field_names: list[str],
    requested_field: str,
) -> str:
    lookup = {
        str(name).strip().lower(): str(name)
        for name in field_names
    }

    matched = lookup.get(
        requested_field.strip().lower()
    )

    if matched is None:
        raise RuntimeError(
            f"Field '{requested_field}' tidak ditemukan. "
            f"Field tersedia: {field_names}"
        )

    return matched


def load_master_ids_from_csv(
    source: Path,
    id_field: str,
) -> set[str]:
    with source.open(
        "r",
        encoding="utf-8-sig",
        newline="",
    ) as stream:
        reader = csv.DictReader(stream)

        if not reader.fieldnames:
            raise RuntimeError(
                f"CSV tidak mempunyai header: {source}"
            )

        actual_field = find_case_insensitive_field(
            list(reader.fieldnames),
            id_field,
        )

        values = {
            normalized
            for row in reader
            if (
                normalized := normalize_quad_id(
                    row.get(actual_field)
                )
            )
        }

    return values


def load_master_ids_from_txt(source: Path) -> set[str]:
    values: set[str] = set()

    for line in source.read_text(
        encoding="utf-8-sig"
    ).splitlines():
        normalized = normalize_quad_id(line)

        if normalized:
            values.add(normalized)

    return values


def locate_excel_header(
    worksheet,
    id_field: str,
) -> tuple[int, int]:
    requested = id_field.strip().lower()

    for row_index, row in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=min(30, worksheet.max_row),
            values_only=True,
        ),
        start=1,
    ):
        for column_index, value in enumerate(
            row,
            start=1,
        ):
            if (
                value is not None
                and str(value).strip().lower() == requested
            ):
                return row_index, column_index

    raise RuntimeError(
        f"Header field '{id_field}' tidak ditemukan "
        f"pada 30 baris pertama sheet '{worksheet.title}'."
    )


def load_master_ids_from_excel(
    source: Path,
    id_field: str,
    sheet_name: str,
) -> set[str]:
    workbook = load_workbook(
        source,
        read_only=True,
        data_only=True,
    )

    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise RuntimeError(
                    f"Sheet '{sheet_name}' tidak ditemukan. "
                    f"Sheet tersedia: {workbook.sheetnames}"
                )

            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook[
                workbook.sheetnames[0]
            ]

        header_row, field_column = locate_excel_header(
            worksheet,
            id_field,
        )

        values: set[str] = set()

        for row in worksheet.iter_rows(
            min_row=header_row + 1,
            min_col=field_column,
            max_col=field_column,
            values_only=True,
        ):
            normalized = normalize_quad_id(row[0])

            if normalized:
                values.add(normalized)

        return values

    finally:
        workbook.close()


def load_master_ids_from_shapefile(
    source: Path,
    id_field: str,
) -> set[str]:
    dbf_path = source.with_suffix(".dbf")

    if not dbf_path.exists():
        raise RuntimeError(
            "Shapefile tidak lengkap. File DBF tidak ditemukan: "
            f"{dbf_path}. Field atribut 'id' berada di file DBF."
        )

    reader = shapefile.Reader(str(source))

    try:
        field_names = [
            field[0]
            for field in reader.fields[1:]
        ]

        actual_field = find_case_insensitive_field(
            field_names,
            id_field,
        )

        field_index = field_names.index(actual_field)

        values: set[str] = set()

        for record in reader.iterRecords():
            normalized = normalize_quad_id(
                record[field_index]
            )

            if normalized:
                values.add(normalized)

        return values

    finally:
        reader.close()


def load_master_quad_ids(
    fallback_records: list[dict],
) -> tuple[set[str], str, bool]:
    """
    Mengembalikan:
    - seluruh master ID;
    - deskripsi sumber;
    - apakah sumber benar-benar master eksternal.

    Jika MASTER_QUADS_SOURCE kosong/tidak ditemukan dan
    MASTER_SOURCE_REQUIRED=false, script memakai ID API aktif sebagai fallback.
    """
    fallback_ids = {
        record["quad_id"]
        for record in fallback_records
    }

    if not CREATE_MASTER_MAP:
        return fallback_ids, "Master map dinonaktifkan", False

    if MASTER_QUADS_SOURCE is None:
        message = (
            "MASTER_QUADS_SOURCE belum diisi. "
            "MASTER_MAP_QUADS hanya memakai ID BBOX aktif."
        )

        if MASTER_SOURCE_REQUIRED:
            raise RuntimeError(message)

        print(f"PERINGATAN: {message}")
        return fallback_ids, "Fallback: Planet API BBOX aktif", False

    if not MASTER_QUADS_SOURCE.exists():
        message = (
            "Sumber master tidak ditemukan: "
            f"{MASTER_QUADS_SOURCE}"
        )

        if MASTER_SOURCE_REQUIRED:
            raise RuntimeError(message)

        print(f"PERINGATAN: {message}")
        print(
            "MASTER_MAP_QUADS memakai ID BBOX aktif sebagai fallback."
        )
        return fallback_ids, "Fallback: Planet API BBOX aktif", False

    suffix = MASTER_QUADS_SOURCE.suffix.lower()

    if suffix == ".csv":
        master_ids = load_master_ids_from_csv(
            MASTER_QUADS_SOURCE,
            MASTER_QUAD_ID_FIELD,
        )
    elif suffix in {".txt", ".list"}:
        master_ids = load_master_ids_from_txt(
            MASTER_QUADS_SOURCE
        )
    elif suffix in {".xlsx", ".xlsm"}:
        master_ids = load_master_ids_from_excel(
            MASTER_QUADS_SOURCE,
            MASTER_QUAD_ID_FIELD,
            MASTER_QUADS_SHEET,
        )
    elif suffix == ".shp":
        master_ids = load_master_ids_from_shapefile(
            MASTER_QUADS_SOURCE,
            MASTER_QUAD_ID_FIELD,
        )
    else:
        raise RuntimeError(
            "Format MASTER_QUADS_SOURCE tidak didukung. "
            "Gunakan .csv, .txt, .xlsx, .xlsm, atau .shp."
        )

    if not master_ids:
        raise RuntimeError(
            f"Tidak ada Quad ID valid pada sumber master: "
            f"{MASTER_QUADS_SOURCE}"
        )

    source_description = str(
        MASTER_QUADS_SOURCE
    )

    return master_ids, source_description, True


# ---------------------------------------------------------------------------
# EXCEL
# ---------------------------------------------------------------------------

def safe_file_uri(path_value: str) -> str | None:
    if not path_value:
        return None

    try:
        return Path(path_value).resolve().as_uri()
    except (OSError, ValueError):
        return None


def backup_excel(excel_path: Path) -> Path | None:
    if not excel_path.exists() or not BACKUP_EXCEL:
        return None

    backup_dir = (
        excel_path.parent
        / "excel_backups"
    )
    backup_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    backup_path = (
        backup_dir
        / (
            f"{excel_path.stem}_"
            f"{timestamp}"
            f"{excel_path.suffix}"
        )
    )

    try:
        shutil.copy2(excel_path, backup_path)
    except PermissionError:
        print(
            "PERINGATAN: backup workbook tidak dibuat karena "
            "file Excel sedang dibuka oleh aplikasi lain."
        )
        return None

    return backup_path


def remove_sheet_if_exists(
    workbook: Workbook,
    sheet_name: str,
) -> None:
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])


def status_fills() -> dict[str, PatternFill]:
    return {
        "Master": PatternFill(
            "solid",
            fgColor="F4B183",
        ),
        "Downloaded": PatternFill(
            "solid",
            fgColor="C6E0B4",
        ),
        "Reused": PatternFill(
            "solid",
            fgColor="BDD7EE",
        ),
        "Existing": PatternFill(
            "solid",
            fgColor="E2F0D9",
        ),
        "Missing": PatternFill(
            "solid",
            fgColor="FFF2CC",
        ),
        "Failed": PatternFill(
            "solid",
            fgColor="F4CCCC",
        ),
        "Extra": PatternFill(
            "solid",
            fgColor="D9D9D9",
        ),
        "Invalid filename": PatternFill(
            "solid",
            fgColor="F4CCCC",
        ),
    }


def style_spatial_map(
    worksheet,
    *,
    all_ids: set[str],
    active_records: list[dict],
    title: str,
    source_description: str,
    master_source_used: bool,
) -> None:
    parsed_ids: dict[tuple[int, int], str] = {}

    for quad_id in sorted(all_ids):
        parsed = parse_quad_id(quad_id)

        if parsed:
            parsed_ids[parsed] = quad_id

    if not parsed_ids:
        raise RuntimeError(
            "Tidak ada ID valid untuk membentuk peta matriks."
        )

    active_by_id = {
        record["quad_id"]: record
        for record in active_records
    }

    x_values = sorted(
        {coordinate[0] for coordinate in parsed_ids}
    )
    y_values = sorted(
        {coordinate[1] for coordinate in parsed_ids},
        reverse=True,
    )

    fills = status_fills()
    title_fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )
    header_fill = PatternFill(
        "solid",
        fgColor="D9EAF7",
    )

    thin_side = Side(
        style="thin",
        color="808080",
    )
    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    last_column = max(2, len(x_values) + 1)
    last_letter = get_column_letter(last_column)

    worksheet.merge_cells(
        f"A1:{last_letter}1"
    )

    title_cell = worksheet["A1"]
    title_cell.value = title
    title_cell.fill = title_fill
    title_cell.font = Font(
        bold=True,
        color="FFFFFF",
        size=14,
    )
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    worksheet.row_dimensions[1].height = 26

    metadata = [
        ("Sumber master", source_description),
        ("Jumlah master quad", len(all_ids)),
        ("Mosaic aktif", MOSAIC_NAME),
        (
            "Mode",
            (
                "Master Indonesia"
                if master_source_used
                else "Fallback BBOX aktif"
            ),
        ),
        (
            "Generated",
            datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
        ),
    ]

    for row_number, (label, value) in enumerate(
        metadata,
        start=2,
    ):
        worksheet.cell(
            row=row_number,
            column=1,
            value=label,
        ).font = Font(bold=True)

        worksheet.cell(
            row=row_number,
            column=2,
            value=value,
        )

    legend_row = 7

    legend = [
        ("Master/Belum diproses", "Master"),
        ("Downloaded", "Downloaded"),
        ("Reused", "Reused"),
        ("Existing", "Existing"),
        ("Missing", "Missing"),
        ("Failed", "Failed"),
        ("Extra", "Extra"),
    ]

    for column_number, (label, status) in enumerate(
        legend,
        start=1,
    ):
        cell = worksheet.cell(
            row=legend_row,
            column=column_number,
            value=label,
        )
        cell.fill = fills[status]
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    header_row = 9

    corner = worksheet.cell(
        row=header_row,
        column=1,
        value="Y \\ X",
    )
    corner.fill = header_fill
    corner.font = Font(bold=True)
    corner.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    corner.border = border

    for excel_column, x_value in enumerate(
        x_values,
        start=2,
    ):
        cell = worksheet.cell(
            row=header_row,
            column=excel_column,
            value=x_value,
        )
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    x_to_column = {
        value: index
        for index, value in enumerate(
            x_values,
            start=2,
        )
    }

    for row_offset, y_value in enumerate(
        y_values,
        start=1,
    ):
        excel_row = header_row + row_offset

        y_cell = worksheet.cell(
            row=excel_row,
            column=1,
            value=y_value,
        )
        y_cell.fill = header_fill
        y_cell.font = Font(bold=True)
        y_cell.alignment = Alignment(horizontal="center")
        y_cell.border = border

        worksheet.row_dimensions[excel_row].height = 23

        for x_value in x_values:
            quad_id = parsed_ids.get(
                (x_value, y_value)
            )

            if quad_id is None:
                continue

            excel_column = x_to_column[x_value]
            cell = worksheet.cell(
                row=excel_row,
                column=excel_column,
                value=quad_id,
            )

            active = active_by_id.get(quad_id)

            if active:
                fill_key = active["status"]
            else:
                fill_key = "Master"

            cell.fill = fills.get(
                fill_key,
                fills["Master"],
            )
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            if active:
                file_uri = safe_file_uri(
                    active["absolute_path"]
                )

                if file_uri:
                    cell.hyperlink = file_uri
                    cell.font = Font(
                        color="0563C1",
                        underline="single",
                    )

    worksheet.column_dimensions["A"].width = 12

    for column_number in range(
        2,
        len(x_values) + 2,
    ):
        worksheet.column_dimensions[
            get_column_letter(column_number)
        ].width = 15

    worksheet.freeze_panes = "B10"
    worksheet.sheet_view.showGridLines = False
    worksheet.auto_filter.ref = None
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_view.zoomScale = 75


def style_detail_sheet(
    worksheet,
    records: list[dict],
    invalid_files: list[Path],
    period_code: str,
) -> None:
    title_fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )
    header_fill = PatternFill(
        "solid",
        fgColor="5B9BD5",
    )
    fills = status_fills()

    thin_side = Side(
        style="thin",
        color="B7B7B7",
    )
    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    headers = [
        "No",
        "Quad ID",
        "X",
        "Y",
        "Mosaic Name",
        "Nama File",
        "Status",
        "Percent Covered",
        "Ukuran File (MB)",
        "Path Relatif",
        "Path Absolut",
        "Waktu Proses",
        "Error",
    ]

    last_letter = get_column_letter(len(headers))

    worksheet.merge_cells(
        f"A1:{last_letter}1"
    )

    title_cell = worksheet["A1"]
    title_cell.value = (
        f"DETAIL QUAD PLANET "
        f"GLOBAL QUARTERLY {period_code}"
    )
    title_cell.fill = title_fill
    title_cell.font = Font(
        bold=True,
        color="FFFFFF",
        size=14,
    )
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet["A2"] = "Mosaic"
    worksheet["A2"].font = Font(bold=True)
    worksheet["B2"] = MOSAIC_NAME

    worksheet["D2"] = "Jumlah"
    worksheet["D2"].font = Font(bold=True)
    worksheet["E2"] = len(records)

    header_row = 4

    for column_number, header in enumerate(
        headers,
        start=1,
    ):
        cell = worksheet.cell(
            row=header_row,
            column=column_number,
            value=header,
        )
        cell.fill = header_fill
        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        cell.border = border

    sorted_records = sorted(
        records,
        key=lambda item: (
            -item["y"],
            item["x"],
        ),
    )

    for number, record in enumerate(
        sorted_records,
        start=1,
    ):
        excel_row = header_row + number

        values = [
            number,
            record["quad_id"],
            record["x"],
            record["y"],
            record["mosaic_name"],
            record["filename"],
            record["status"],
            record["coverage"],
            record["size_mb"],
            record["relative_path"],
            record["absolute_path"],
            record["processed_at"],
            record["error"],
        ]

        for column_number, value in enumerate(
            values,
            start=1,
        ):
            cell = worksheet.cell(
                row=excel_row,
                column=column_number,
                value=value,
            )
            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        worksheet.cell(
            row=excel_row,
            column=7,
        ).fill = fills[record["status"]]

        file_uri = safe_file_uri(
            record["absolute_path"]
        )

        if file_uri:
            filename_cell = worksheet.cell(
                row=excel_row,
                column=6,
            )
            filename_cell.hyperlink = file_uri
            filename_cell.font = Font(
                color="0563C1",
                underline="single",
            )

        worksheet.cell(
            row=excel_row,
            column=9,
        ).number_format = "0.00"

        worksheet.cell(
            row=excel_row,
            column=12,
        ).number_format = "yyyy-mm-dd hh:mm:ss"

    next_number = len(sorted_records) + 1

    for invalid_path in invalid_files:
        excel_row = header_row + next_number

        values = [
            next_number,
            "",
            "",
            "",
            MOSAIC_NAME,
            invalid_path.name,
            "Invalid filename",
            "",
            round(
                invalid_path.stat().st_size
                / (1024 * 1024),
                2,
            ),
            str(invalid_path),
            str(invalid_path.resolve()),
            datetime.now(),
            "Nama TIFF tidak mengikuti pola angka-angka.tif",
        ]

        for column_number, value in enumerate(
            values,
            start=1,
        ):
            cell = worksheet.cell(
                row=excel_row,
                column=column_number,
                value=value,
            )
            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        worksheet.cell(
            row=excel_row,
            column=7,
        ).fill = fills["Invalid filename"]

        next_number += 1

    final_row = max(
        header_row
        + len(sorted_records)
        + len(invalid_files),
        header_row,
    )

    worksheet.auto_filter.ref = (
        f"A{header_row}:{last_letter}{final_row}"
    )
    worksheet.freeze_panes = "A5"
    worksheet.sheet_view.showGridLines = False

    widths = {
        "A": 7,
        "B": 15,
        "C": 9,
        "D": 9,
        "E": 34,
        "F": 22,
        "G": 18,
        "H": 18,
        "I": 18,
        "J": 42,
        "K": 55,
        "L": 21,
        "M": 42,
    }

    for column_letter, width in widths.items():
        worksheet.column_dimensions[
            column_letter
        ].width = width


def save_workbook_safely(
    workbook: Workbook,
    target_path: Path,
) -> tuple[Path, bool]:
    """
    Menyimpan workbook utama.

    Jika file target sedang dibuka di Excel pada Windows, openpyxl
    akan mendapat PermissionError. Dalam kondisi tersebut proses
    tidak digagalkan; workbook disimpan sebagai file recovery.
    """
    try:
        temporary_path = target_path.with_name(
            f".{target_path.stem}.tmp{target_path.suffix}"
        )
        workbook.save(temporary_path)
        temporary_path.replace(target_path)
        return target_path, False

    except PermissionError:
        timestamp = datetime.now().strftime(
            "%Y%m%d_%H%M%S"
        )

        recovery_path = (
            target_path.parent
            / (
                f"{target_path.stem}_RECOVERY_"
                f"{timestamp}{target_path.suffix}"
            )
        )

        workbook.save(recovery_path)

        print()
        print("=" * 74)
        print("PERINGATAN FILE EXCEL TERKUNCI")
        print("=" * 74)
        print(
            "Workbook utama tidak dapat ditimpa karena kemungkinan "
            "sedang dibuka di Microsoft Excel."
        )
        print(f"Target utama : {target_path}")
        print(f"File recovery: {recovery_path}")
        print(
            "Tutup workbook utama sebelum run berikutnya agar sheet "
            "dapat diperbarui langsung pada file utama."
        )

        return recovery_path, True


def export_quads_excel(
    records: list[dict],
    invalid_files: list[Path],
    excel_path: Path,
    run_name: str,
) -> Path:
    excel_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()
    workbook.remove(workbook.active)

    period_code = extract_period_code(MOSAIC_NAME)

    map_sheet_name = safe_sheet_name(
        f"MAP_{period_code}",
        "MAP",
    )
    detail_sheet_name = safe_sheet_name(
        f"DETAIL_{period_code}",
        "DETAIL",
    )

    # Bersihkan sheet master versi lama agar tidak membingungkan.
    remove_sheet_if_exists(
        workbook,
        "MASTER_QUAD_IDS",
    )

    if CREATE_MASTER_MAP:
        master_ids, source_description, master_used = (
            load_master_quad_ids(records)
        )

        remove_sheet_if_exists(
            workbook,
            MASTER_MAP_SHEET_NAME,
        )

        master_sheet = workbook.create_sheet(
            MASTER_MAP_SHEET_NAME,
            0,
        )

        style_spatial_map(
            master_sheet,
            all_ids=master_ids,
            active_records=records,
            title=(
                "PETA POSISI SELURUH QUAD "
                "PLANET INDONESIA"
            ),
            source_description=source_description,
            master_source_used=master_used,
        )

    remove_sheet_if_exists(
        workbook,
        map_sheet_name,
    )
    remove_sheet_if_exists(
        workbook,
        detail_sheet_name,
    )

    active_ids = {
        record["quad_id"]
        for record in records
    }

    active_map_sheet = workbook.create_sheet(
        map_sheet_name
    )

    style_spatial_map(
        active_map_sheet,
        all_ids=active_ids,
        active_records=records,
        title=(
            f"PETA POSISI QUAD PLANET "
            f"GLOBAL QUARTERLY {period_code}"
        ),
        source_description="Planet API sesuai BBOX aktif",
        master_source_used=False,
    )

    detail_sheet = workbook.create_sheet(
        detail_sheet_name
    )

    style_detail_sheet(
        detail_sheet,
        records,
        invalid_files,
        period_code,
    )

    saved_path, used_recovery = save_workbook_safely(
        workbook,
        excel_path,
    )

    if used_recovery:
        print(
            "STATUS EXCEL : RECOVERY "
            "(workbook utama sedang terkunci)"
        )
    else:
        print("STATUS EXCEL : UPDATED")

    return saved_path


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main() -> None:
    if not API_KEY:
        raise RuntimeError(
            "PL_API_KEY belum diisi di file .env."
        )

    if not MOSAIC_NAME:
        raise RuntimeError(
            "MOSAIC_NAME belum diisi di file .env."
        )

    if not BBOX:
        raise RuntimeError(
            "BBOX belum diisi di file .env."
        )

    bbox = validate_bbox(BBOX)
    started_at = datetime.now()
    run_context = create_run_context(
        started_at
    )
    run_name = str(
        run_context["run_name"]
    )
    output_folder = Path(
        run_context["output_folder"]
    )
    requested_excel_path = Path(
        run_context["excel_path"]
    )

    write_run_state(
        run_context,
        status="running",
    )

    print("=" * 74)
    print("PLANET GLOBAL QUARTERLY QUAD DOWNLOADER")
    print("=" * 74)
    print(f"Project root       : {BASE_DIR}")
    print(f"Run name           : {run_name}")
    print(f"Mosaic             : {MOSAIC_NAME}")
    print(f"BBOX               : {bbox}")
    print(f"Workers            : {MAX_WORKERS}")
    print(f"Dry run            : {DRY_RUN}")
    print(f"Output folder      : {output_folder}")
    print(f"Excel output       : {requested_excel_path}")
    print(f"Reuse TIFF         : {REUSE_PREVIOUS_TIFF}")
    print(f"Reuse method       : {REUSE_METHOD}")
    print(f"Create master map  : {CREATE_MASTER_MAP}")
    print(
        f"Master source      : "
        f"{MASTER_QUADS_SOURCE or '(belum diatur)'}"
    )
    print(f"Use selected IDs   : {USE_SELECTED_QUADS}")
    if USE_SELECTED_QUADS:
        print(f"Selected file      : {SELECTED_QUADS_FILE}")
    print(f"RUN_NAME={run_name}")
    print(f"RUN_FOLDER={output_folder.resolve()}")
    print(f"RUN_EXCEL={requested_excel_path.resolve()}")
    print()

    try:
        mosaic = find_mosaic()
        mosaic_id = str(mosaic["id"])

        print(f"Mosaic ID          : {mosaic_id}")

        quads = list_quads(mosaic_id, bbox)

        api_quad_count = len(quads)
        selected_ids = load_selected_quad_ids()

        if selected_ids is not None:
            api_ids = {
                str(quad.get("id"))
                for quad in quads
                if quad.get("id")
            }

            quads = [
                quad
                for quad in quads
                if str(quad.get("id")) in selected_ids
            ]

            unavailable_ids = sorted(
                selected_ids - api_ids
            )

            print()
            print(f"Quad hasil BBOX API   : {api_quad_count}")
            print(f"Quad pilihan UI       : {len(selected_ids)}")
            print(f"Quad setelah filter   : {len(quads)}")
            print(f"File filter           : {SELECTED_QUADS_FILE}")

            if unavailable_ids:
                preview = ", ".join(unavailable_ids[:10])
                suffix = (
                    " ..."
                    if len(unavailable_ids) > 10
                    else ""
                )
                print(
                    "PERINGATAN: "
                    f"{len(unavailable_ids)} ID pilihan tidak ditemukan "
                    f"pada respons API: {preview}{suffix}"
                )
        else:
            print()
            print(f"Quad ditemukan: {len(quads)}")

        if not quads:
            write_run_state(
                run_context,
                status="empty",
                completed_at=datetime.now(),
                counts={
                    "downloaded": 0,
                    "reused": 0,
                    "existing": 0,
                    "failed": 0,
                },
                error="Tidak ada quad pada BBOX tersebut.",
            )
            print("Tidak ada quad pada BBOX tersebut.")
            return

        previous_tiff_index = build_previous_tiff_index(
            output_folder
        )

        if REUSE_PREVIOUS_TIFF:
            print(
                "TIFF reusable ditemukan: "
                f"{len(previous_tiff_index)}"
            )

        download_results: dict[
            str,
            dict[str, str],
        ] = {}

        downloaded = 0
        reused = 0
        existing = 0
        failed = 0

        if DRY_RUN:
            print()
            print(
                "DRY_RUN=true: download tidak dijalankan. "
                "Excel dibuat berdasarkan file pada folder run ini."
            )
        else:
            print()
            print("Memulai download TIFF...")

            with ThreadPoolExecutor(
                max_workers=MAX_WORKERS
            ) as executor:
                futures = {
                    executor.submit(
                        download_quad,
                        quad,
                        output_folder,
                        previous_tiff_index,
                    ): str(quad["id"])
                    for quad in quads
                }

                for future in tqdm(
                    as_completed(futures),
                    total=len(futures),
                    unit="quad",
                    desc="Processing",
                ):
                    quad_id = futures[future]

                    try:
                        result = future.result()

                        download_results[quad_id] = {
                            "status": result,
                            "error": "",
                        }

                        if result == "existing":
                            existing += 1
                        elif result.startswith("reused_"):
                            reused += 1
                        else:
                            downloaded += 1

                    except Exception as exc:
                        failed += 1

                        download_results[quad_id] = {
                            "status": "failed",
                            "error": str(exc),
                        }

                        print(
                            f"\nGagal memproses {quad_id}: {exc}"
                        )

        print()
        print("=" * 74)
        print("PROSES TIFF SELESAI")
        print("=" * 74)
        print(f"Downloaded : {downloaded}")
        print(f"Reused     : {reused}")
        print(f"Existing   : {existing}")
        print(f"Failed     : {failed}")
        print(f"Folder     : {output_folder}")

        records, invalid_files = build_active_records(
            quads,
            download_results,
            output_folder,
        )

        status_counts: dict[str, int] = {}

        for record in records:
            status = record["status"]
            status_counts[status] = (
                status_counts.get(status, 0) + 1
            )

        print()
        print("STATUS FILE:")

        for status, count in sorted(
            status_counts.items()
        ):
            print(f"- {status}: {count}")

        saved_excel_path = requested_excel_path

        if EXPORT_EXCEL:
            print()
            print("Membuat workbook Excel per sesi...")

            saved_excel_path = export_quads_excel(
                records,
                invalid_files,
                requested_excel_path,
                run_name,
            )

            period_code = extract_period_code(
                MOSAIC_NAME
            )

            print(f"Excel berhasil     : {saved_excel_path}")

            if CREATE_MASTER_MAP:
                print(
                    f"Sheet pertama      : "
                    f"{MASTER_MAP_SHEET_NAME}"
                )

            print(
                f"Sheet periode      : "
                f"MAP_{period_code}, "
                f"DETAIL_{period_code}"
            )

        counts = {
            "downloaded": downloaded,
            "reused": reused,
            "existing": existing,
            "failed": failed,
            "total": len(quads),
        }

        write_run_state(
            run_context,
            status=(
                "completed"
                if failed == 0
                else "completed_with_errors"
            ),
            completed_at=datetime.now(),
            counts=counts,
            excel_path=saved_excel_path,
        )

        print()
        print("=" * 74)
        print("RUN SELESAI")
        print("=" * 74)
        print(f"RUN_NAME={run_name}")
        print(f"RUN_FOLDER={output_folder.resolve()}")
        print(f"RUN_EXCEL={saved_excel_path.resolve()}")
        print(f"RUN_MANIFEST={Path(run_context['manifest_path']).resolve()}")

    except Exception as exc:
        write_run_state(
            run_context,
            status="failed",
            completed_at=datetime.now(),
            counts={},
            error=str(exc),
        )
        raise


